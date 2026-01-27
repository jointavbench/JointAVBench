import json
import os
import pickle
from collections import defaultdict
from datetime import datetime, timedelta
from scenedetect import detect, ContentDetector, open_video, SceneManager
import cv2
import re
import pandas as pd
from openpyxl import load_workbook


def xlsx_to_json(file_path, sheet_name=None, header_row=1, output_file=None):
    """
    Convert Excel file (.xlsx) to JSON format.

    Args:
        file_path (str): Path to Excel file
        sheet_name (str): Worksheet name, if None uses the first worksheet
        header_row (int): Row number containing headers (starts from 1)
        output_file (str): Output JSON file path, if None doesn't save to file

    Returns:
        list: List of dictionaries containing all data in JSON format
    """
    # Load workbook
    wb = load_workbook(filename=file_path, read_only=True)

    # Get worksheet
    if sheet_name is None:
        sheet = wb.active
    else:
        sheet = wb[sheet_name]

    # Extract headers
    headers = []
    for cell in sheet[header_row]:
        headers.append(cell.value)

    # Collect data rows
    data = []
    for row in sheet.iter_rows(min_row=header_row + 1):
        row_data = {}
        for idx, cell in enumerate(row):
            # Ensure we don't exceed header range
            if idx < len(headers):
                row_data[headers[idx]] = cell.value
        data.append(row_data)

    # Close workbook
    wb.close()

    # Save to file if requested
    if output_file:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

    return data

def load_json(file_path):
    """Load JSON file and return parsed data."""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def load_jsonl(file_path):
    """
    Load JSONL (JSON Lines) file where each line is a separate JSON object.

    Args:
        file_path (str): Path to JSONL file

    Returns:
        list: List of parsed JSON objects
    """
    data_list = []
    with open(file_path, 'r', encoding='utf-8') as file:
        for line in file:
            # Parse each line as a JSON object
            json_obj = json.loads(line.strip())
            data_list.append(json_obj)
    return data_list

def save_json(data, file_path):
    """Save data to JSON file with pretty printing."""
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def jsonl2xlsx(file_path, save_path):
    """Convert JSONL file to Excel format."""
    data = load_jsonl(file_path)
    df = pd.DataFrame(data)
    df.to_excel(save_path, index=False)

def json2xlsx(file_path, save_path):
    """Convert JSON file to Excel format."""
    data = load_json(file_path)
    df = pd.DataFrame(data)
    df.to_excel(save_path, index=False)

def save_xlsx(data, save_path):
    """Save data to Excel file."""
    df = pd.DataFrame(data)
    df.to_excel(save_path, index=False)

def load_caption(file_path):
    """
    Load captions and organize by video name.

    Args:
        file_path (str): Path to caption JSON file

    Returns:
        dict: Dictionary mapping video_name to list of caption segments
    """
    data = load_json(file_path)
    transformed_data = defaultdict(list)
    for item in data:
        video_name = item['video_name']
        transformed_data[video_name].append(item)
    return transformed_data

def load_desc(file_path):
    """
    Load video descriptions and organize by video name.

    Args:
        file_path (str): Path to descriptions JSON file

    Returns:
        dict: Dictionary mapping video_name to description text
    """
    descs = load_json(file_path)
    return dict({desc['video_name']: desc['description'] for desc in descs})

def load_sub(sub_dir):
    """
    Load subtitle files from directory.

    Args:
        sub_dir (str): Directory containing subtitle JSON files

    Returns:
        dict: Dictionary mapping video_name to list of subtitle segments
    """
    files = os.listdir(sub_dir)
    vid2sub = dict()
    for file in files:
        transcription = load_json(os.path.join(sub_dir, file))
        segments = transcription['segments']
        vid_name = file.split('.')[0]
        vid2sub[vid_name] = segments
    return vid2sub

def seconds2timestamp(seconds):
    """
    Convert seconds to timestamp format HH:MM:SS.mmm.

    Args:
        seconds (float): Time in seconds

    Returns:
        str: Formatted timestamp string
    """
    t = timedelta(seconds=seconds)

    total_seconds = int(t.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    secs = total_seconds % 60
    milliseconds = t.microseconds // 1000

    time_format = f"{hours:02}:{minutes:02}:{secs:02}.{milliseconds:03}"
    return time_format

def timestamp2seconds(timestamp):
    """
    Convert timestamp string to seconds.

    Args:
        timestamp (str): Timestamp in format "HH:MM:SS.ffffff"

    Returns:
        float: Time in seconds
    """
    time_obj = datetime.strptime(timestamp, "%H:%M:%S.%f")
    delta = timedelta(hours=time_obj.hour, minutes=time_obj.minute,
                     seconds=time_obj.second, microseconds=time_obj.microsecond)
    return delta.total_seconds()

def transform_options(item):
    """
    Add letter prefixes (A., B., C., D.) to multiple choice options.

    Args:
        item (dict): Dictionary containing 'options' and 'correct_answer' keys

    Returns:
        dict: Copy of item with prefixed options and correct_prefix field added
    """
    if not isinstance(item, dict) or 'options' not in item or 'correct_answer' not in item:
        return item

    # Generate option prefixes (A., B., C., D., ...)
    prefixes = [f"{chr(65 + i)}." for i in range(len(item['options']))]

    # Create prefixed option list
    prefixed_options = [f"{prefix} {option}"
                       for prefix, option in zip(prefixes, item['options'])]

    correct_index = item['options'].index(item['correct_answer'])
    correct_prefix = prefixes[correct_index]

    # Update dictionary
    transformed_item = item.copy()
    transformed_item['options'] = prefixed_options
    transformed_item['correct_prefix'] = correct_prefix

    return transformed_item

def remove_tone(text):
    """
    Remove phrases like 'with a happy tone', 'with an angry tone' from text.

    Uses regex pattern to match and remove tone descriptions.

    Args:
        text (str): Input text

    Returns:
        str: Text with tone phrases removed
    """
    pattern = r'\bwith an?\s+(\S+\s+){1,3}tone\b'
    result = re.sub(pattern, '', text)
    result = ' '.join(result.split())
    return result

def prepare_sub(subtitles, video_captions, include_timestamp = True):
    if include_timestamp:
        for vc in video_captions:
            vc_start = timestamp2seconds(vc['start_time'])
            vc_end = timestamp2seconds(vc['end_time'])
            vc['subtitle'] = ""

            for subtitle in subtitles:
                if not (subtitle['start'] >= vc_end or subtitle['end'] <= vc_start):
                    vc['subtitle'] += f"[{seconds2timestamp(subtitle['start']-vc_start)}-{seconds2timestamp(subtitle['end']-vc_start)}]{subtitle['text']}\n"
    else:
        for vc in video_captions:
            vc_start = timestamp2seconds(vc['start_time']) if isinstance(vc['start_time'], str) else vc['start_time']
            vc_end = timestamp2seconds(vc['end_time']) if isinstance(vc['end_time'], str) else vc['end_time']
            vc['subtitle'] = ""

            for subtitle in subtitles:
                if not (subtitle['start'] >= vc_end or subtitle['end'] <= vc_start):
                    vc['subtitle'] += f"{subtitle['text']}\n"
    return video_captions

def join_subtitles(subtitles, start_time = 0):
    joined_sub = ''
    for subtitle in subtitles:
        joined_sub += f"From {max(subtitle['start']-start_time,0):.2f} to {subtitle['end']-start_time:.2f}: {subtitle['text'].strip()}\n"
    return joined_sub

def merge_captions_segment_wise(video_captions, subtitles, aud_forms, segment_id=None, vid_desc=None, use_vc = True):
    all_caps = ''
    if vid_desc is not None:
        all_caps += f"Story Description: {vid_desc}\n"
    
    if segment_id:
        if isinstance(segment_id, int):
            # Process only the specific segment_id
            target_segment_id = segment_id
            for vc in video_captions:
                if vc['segment_id'] == target_segment_id:
                    vc_start_time = timestamp2seconds(vc['start_time'])
                    vc_end_time = timestamp2seconds(vc['end_time'])
                    all_caps += f"From {vc_start_time-vc_start_time:.2f} to {vc_end_time-vc_start_time:.2f}:\n"
                    if use_vc:
                        all_caps += f"Video Caption: {vc['video_caption']}\n" if vc['video_caption'] != '' else ''
                    if 'speech' in aud_forms:
                        segment_subtitle = join_subtitles([sub for sub in subtitles if not (sub['end'] <= vc_start_time or sub['start'] >= vc_end_time)], vc_start_time)
                        all_caps += f"Subtitles: {segment_subtitle}\n" if segment_subtitle != '' else ''
                    if 'speech_emotion' in aud_forms:
                        all_caps += f"Speech Emotion: {vc['speech_emotion']}\n" if vc['speech_emotion'] != '' else ''
                    if 'sound_event' in aud_forms:
                        all_caps += f"Sound Event: {vc['sound_event']}\n" if vc['sound_event'] != '' else ''
                    if 'music' in aud_forms:
                        all_caps += f"Music: {vc['music']}\n" if vc['music'] != '' else ''
                    break  # Exit loop after processing the target segment
        elif isinstance(segment_id, list) and len(segment_id) == 2:
            # Process from the first segment_id to the second segment_id
            start_segment_id, end_segment_id = min(segment_id), max(segment_id)
            for vc in video_captions:
                if start_segment_id <= vc['segment_id'] <= end_segment_id:
                    all_caps += f"[Segment {vc['segment_id']}] From {timestamp2seconds(vc['start_time']):.2f} to {timestamp2seconds(vc['end_time']):.2f}:\n"
                    if use_vc:
                        all_caps += f"Video Caption: {vc['video_caption']}\n" if vc['video_caption'] != '' else ''
                    if 'speech' in aud_forms:
                        segment_subtitle = join_subtitles([sub for sub in subtitles if not (sub['end'] <= timestamp2seconds(vc['start_time']) or sub['start'] >= timestamp2seconds(vc['end_time']))])
                        all_caps += f"Subtitles: {segment_subtitle}\n" if segment_subtitle != '' else ''
                    if 'speech_emotion' in aud_forms:
                        all_caps += f"Speech Emotion: {vc['speech_emotion']}\n" if vc['speech_emotion'] != '' else ''
                    if 'sound_event' in aud_forms:
                        all_caps += f"Sound Event: {vc['sound_event']}\n" if vc['sound_event'] != '' else ''
                    if 'music' in aud_forms:
                        all_caps += f"Music: {vc['music']}\n" if vc['music'] != '' else ''
    else:
        # Process all video_captions
        for vc in video_captions:
            all_caps += f"[Segment {vc['segment_id']}] From {timestamp2seconds(vc['start_time']):.2f} to {timestamp2seconds(vc['end_time']):.2f}:\n"
            if use_vc: 
                all_caps += f"Video Caption: {vc['video_caption']}\n" if vc['video_caption'] != '' else ''
            if 'speech' in aud_forms:
                segment_subtitle = join_subtitles([sub for sub in subtitles if not (sub['end'] <= timestamp2seconds(vc['start_time']) or sub['start'] >= timestamp2seconds(vc['end_time']))])
                all_caps += f"Subtitles: {segment_subtitle}\n" if segment_subtitle != '' else ''
            if 'speech_emotion' in aud_forms:
                all_caps += f"Speech Emotion: {vc['speech_emotion']}\n" if vc['speech_emotion'] != '' else ''
            if 'sound_event' in aud_forms:
                all_caps += f"Sound Event: {vc['sound_event']}\n" if vc['sound_event'] != '' else ''
            if 'music' in aud_forms:
                all_caps += f"Music: {vc['music']}\n" if vc['music'] != '' else ''
    
    # all_caps += f'Subtitles:\n{subtitles}'
    return all_caps

def join_captions(video_captions):
    vid_caption = ''
    music_caption = ''
    sound_caption = ''
    speech_emotion = ''
    for vc in video_captions:
        vid_caption += f"[Segment {vc['segment_id']}] From {timestamp2seconds(vc['start_time']):.2f} to {timestamp2seconds(vc['end_time']):.2f}: {vc['video_caption']}\n" if vc['video_caption'] != '' else ''
        music_caption += f"[Segment {vc['segment_id']}] From {timestamp2seconds(vc['start_time']):.2f} to {timestamp2seconds(vc['end_time']):.2f}: {vc['music']}\n" if vc['music'] != '' else ''
        sound_caption += f"[Segment {vc['segment_id']}] From {timestamp2seconds(vc['start_time']):.2f} to {timestamp2seconds(vc['end_time']):.2f}: {vc['sound_event']}\n" if vc['sound_event'] != '' else ''
        speech_emotion += f"[Segment {vc['segment_id']}] From {timestamp2seconds(vc['start_time']):.2f} to {timestamp2seconds(vc['end_time']):.2f}: {vc['speech_emotion']}\n" if vc['speech_emotion'] != '' else ''
    return vid_caption, music_caption, sound_caption, speech_emotion

def pair_captions(video_captions, audio_captions, ser_captions):
    matched_results = []
    
    audio_caption_dict = {
        (item['video_name'], item['segment_id']): item for item in audio_captions
    }
    ser_caption_dict = {
        (item['video_name'], item['segment_id']): item for item in ser_captions
    }
    # 遍历video_captions，查找匹配的audio_caption
    for video_item in video_captions:
        key = (video_item['video_name'], video_item['segment_id'])
        matched_results.append({
            'video_name': video_item['video_name'],
            'segment_id': video_item['segment_id'],
            'start_time': video_item['start_time'],
            'end_time': video_item['end_time'],
            'video_caption': video_item['caption'],
            'audio_caption': audio_caption_dict[key]['caption'] if key in audio_caption_dict else "",
            'speech_emotion': ser_caption_dict[key]['caption'] if key in ser_caption_dict else ""
        })
    
    return matched_results

if __name__ == '__main__':
    ser_caption = load_json('./raw_data/ser_caption_v1.json')
    audio_caption = load_json('./raw_data/audio_caption_v1.json')
    video_caption = load_json('./raw_data/video_caption.json')
    matched_results = pair_captions(video_caption, audio_caption, ser_caption)
    save_json(matched_results, './raw_data/paired_captions_v1.json')